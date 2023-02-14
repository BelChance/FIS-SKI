import openpyxl                                # python与excel交互
from pyecharts import options as opts          # 导入pyecharts库
from pyecharts.charts import Bar               # 柱状图
from pyecharts.charts import Line              # 折线图

# 导入工作簿---excel需要为xlsx格式
workbook = openpyxl.load_workbook("cross_country_stat(越野滑雪).xlsx")

# 导入最终处理后的final工作表
final = workbook['final']

# 获取该工作表的行号和列号
row = final.max_row             # row=1822
column = final.max_column       # column=8
# print(row, column)

'''# 依次得出每列的名称
final_head = final["A1":"H1"]
for r in final_head:
    for c in r:
        # print(c.value)
# 列名依次是year,place,nation, dics,gender,nation1,nation2,nation3'''

# 依次得出每列的名称
for i in range(1, column + 1):
    head = final.cell(1, i).value
    # print(head.value)
# 列名依次是year,place,nation, dics,gender,nation1,nation2,nation3


# 定义计数函数--列表转化字典
def count(List):
    dict = {}
    for area in List:
        keys = area.split(",")
        for key in keys:
            if key in dict.keys():
                dict[key] = dict[key]+1
            else:
                dict[key] = 1
    return dict


# 用列表存放数据
list_gender = []
list_dics = []
list_nation1 = []
list_nation2 = []
list_nation3 = []

# 获取性别列，并转化为字典
for i in range(2, row + 1):
    s = final.cell(i, 5).value
    list_gender.append(s)
dict_gender = count(list_gender)

# 获取事件列,并转化为字典
for i in range(2, row + 1):
    s = final.cell(i, 4).value
    list_dics.append(s)
dict_dics = count(list_dics)

# 获取nation1列，并转化为字典
for i in range(2, row + 1):
    s = final.cell(i, 6).value
    list_nation1.append(s)
dict_nation1 = count(list_nation1)

# 获取rank2列，并转化为字典
for i in range(2, row + 1):
    s = final.cell(i, 7).value
    list_nation2.append(s)
dict_nation2 = count(list_nation2)

# 获取rank3列，并转化为字典
for i in range(2, row + 1):
    s = final.cell(i, 8).value
    list_nation3.append(s)
dict_nation3 = count(list_nation3)

# 根据事件类型dics做出柱状图
bar = (
    Bar()
    .add_xaxis(list(dict_dics.keys()))
    .add_yaxis("事件类型", list(dict_dics.values()))
    .set_global_opts(
        xaxis_opts=opts.AxisOpts(axislabel_opts=opts.LabelOpts(rotate=-15)),
        title_opts=opts.TitleOpts(title="dics图表分析"),
    )
)
bar.render("dics.html")

# 根据性别gender做出柱状图
bar = (
    Bar()
    .add_xaxis(list(dict_gender.keys()))
    .add_yaxis("gender", list(dict_gender.values()))
    .set_global_opts(
        xaxis_opts=opts.AxisOpts(axislabel_opts=opts.LabelOpts(rotate=-15)),
        title_opts=opts.TitleOpts(title="gender图表分析"),
    )
)
bar.render("gender.html")


# 根据国家排名nation1，nation2，nation3做出折线图
line = (
    Line()
    .add_xaxis(list(dict_nation1.keys()))
    .add_yaxis("rank1", list(dict_nation1.values()), is_connect_nones=True)
    .set_global_opts(title_opts=opts.TitleOpts(title="国家排名rank"))
    .render("nation1.html")
)

line = (
    Line()
    .add_xaxis(list(dict_nation2.keys()))
    .add_yaxis("rank2", list(dict_nation2.values()), is_connect_nones=True)
    .set_global_opts(title_opts=opts.TitleOpts(title="国家排名rank"))
    .render("nation2.html")
)

line = (
    Line()
    .add_xaxis(list(dict_nation3.keys()))
    .add_yaxis("rank3", list(dict_nation3.values()), is_connect_nones=True)
    .set_global_opts(title_opts=opts.TitleOpts(title="国家排名rank"))
    .render("nation3.html")
)
